import openpyxl
import pymongo
from pymongo import MongoClient

wb = openpyxl.load_workbook('data.xlsx')

sheet = wb['Sheet1']

client = MongoClient()

db = client.Stock_Information

collection = db.news
a = 0

for row in range(2, sheet.max_row + 1):
    a = a + 1
    print (a)
    Stock = str((sheet.cell(row, 2).value))
    Sector = str((sheet.cell(row, 3).value))
    News_Headline = ((sheet.cell(row, 4).value))
    News_fullText = ((sheet.cell(row, 5).value))
    Time = str((sheet.cell(row, 6).value))
    links = str((sheet.cell(row, 7).value))

    links = links.split(",")
    no_of_links = (len(links))
    linkss = []
    for i in range(no_of_links):
        linkss.append(links[i])

    sentiment = str((sheet.cell(row, 8).value))
    if (Stock == "None"):
        Stock = ''

    data = {

        '_id': a,
        "Stock": Stock,
        "Sector": Sector,
        "News_Headline": News_Headline,
        "News_fullText": News_fullText,
        "Time": Time,
        "Sources": linkss,

        "Sentimet": sentiment
    }

    collection.insert(data)

